
from openpyxl import load_workbook

#打开工作薄  定位表单
wb = load_workbook('python9.xlsx')
sheet = wb['python9']

all_data = []
for j in range (1,sheet.max_row+1):#控制行
    row_data = []
    for i in range(1,sheet.max_column+1):#控制列
        row_data.append(sheet.cell(j,i).value)
        all_data.append(row_data)
print(all_data)
