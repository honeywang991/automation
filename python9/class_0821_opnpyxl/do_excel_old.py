# 把python9.xlsx中的python9第一列数据全部读出来，并存到一个列表
from openpyxl import load_workbook

#打开工作薄  定位表单
wb = load_workbook('python9.xlsx')
sheet = wb['python9']

#获取值

#第一个小题目：把第一行的数据  全部获取出来存到一个列表里面
#方法一  ： 一个一个的去读
# test_data = []
# test_data.append(sheet.cell(1,1).value)
# test_data.append(sheet.cell(1,2).value)
# test_data.append(sheet.cell(1,3).value)
# test_data.append(sheet.cell(1,4).value)
# test_data.append(sheet.cell(1,5).value)
# print(test_data)

#
# #方法二： 控制列表循环  利用for循环
# test_data = []
# #经过方法一  我们观察到  列是变化的 1 2 3  4 5
# #rang(1,6)  可以生成 1 2 3 4 5
#
# print(sheet.max_column)#获取有数据的最大列
# print(sheet.max_row)#获取有数据的最大行
#
#
# for i in range(1,sheet.max_column+1):
#     test_data.append(sheet.cell(1,i).value)
# print(test_data)
#
# print('==================')
#
# col_data = []
# #第二题： 把第一列的所有数据读出来存到一个列表里面
# for i in range(1,sheet.max_row+1):
#     col_data.append(sheet.cell(i,6).value)
# print(col_data)

#题目三：所有数据全部独出来，每一行存在一个自列表里面
#所有行的数据存到一个大列表里面
#嵌套列表

#分解步骤
#1：先完成第一行数据的读取
#2：再来想办法 完成第二行数据的读取
# all_data = []
# for j in range (1,sheet.max_row+1):#控制行
#     row_data = []
#     for i in range(1,sheet.max_column+1):#控制列
#         row_data.append(sheet.cell(j,i).value)
#         all_data.append(row_data)
# print(all_data)

#数据从Excel里面读取出来是什么类型？有没有关注过？

# print(sheet.cell(3,1).value)
print(sheet.cell(3,6).value)
print(type(sheet.cell(3,6).value))
# print(type(eval(sheet.cell(3,6).value)))
#除了 int float 是原来的数据格式，其他都是str

