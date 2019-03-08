#引入模块 import  from ... import
#一层层剥开，放在lib下面  我不管
from class_0821_opnpyxl.test_python9.test_1 import Add
#命名规范:数字 字母 下划线组成 不能以数字开头
#类名  首字母大写驼峰命名
#函数 模块名 变量名 小写 不同字母用下划线隔开  方便阅读
#解决我们测试数据的问题
#测试用例：数据库 Excel
#怎么操作Excel
#打开工作薄--找到表单--定位到你要的单元格
#openpyxl 专门操作Excel 同时读写  pip install  openpyxl
#  xlrd:只能读    xlwt:只能写
#load_workbook  打开excel
#Workbook 新建一个工作薄
from openpyxl import load_workbook

#1:打开工作薄，必须是已经存在的工作薄
#openpyxl 操作excel 必须是xlsx结尾   不能是xls，否则会报错
wb=load_workbook('t1.xlsx')  #大小写敏感
#坑一： 新建文件在windows桌面建好，粘贴进项目
#新建excel不可以在pycharm中新建，在window文件夹中新建
#或者 在桌面新建好，复制到pycharm中
#尽量用绝对路径，换电脑的问题

#2:定位表单  sheet1  sheet2 sheet3
sheet = wb['python9']#最新定义的一个用法
# sheet = wb.get_sheet_by_name('python9')#已经过时

#3：定位单元格，获取单元格里面的数据
#根据行列去定位  坐标 去定位元素
print(sheet.cell(2,3).value)#读取数据
#如果你获取的值   返回的是None 没有数据  要么就是没有取到
sheet.cell(3,3).value='菜菜'#

#如果有数据更新，那么就一定要记得保存文件
wb.save('t1.xlsx')

#坑二：
#如果涉及到写入数据，那么就一定要关闭，Excel一定不能是打开状态？
#不可以在C盘 下面搞代码，因为需要管理员权限
