
from openpyxl import load_workbook
#请把每一行数据存到字典里面，其中key为对于的表头，所有字典存到一个大列表里面

class DoExcel:
    def __init__(self,file_name,sheet_name):#Excel文件名  表单名
        self.file_name=file_name
        self.sheet_name=sheet_name

    def get_title(self):#获取第一行表头字段
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        title = []  #存储表头
        for i in range(1,sheet.max_column+1):
            title.append(sheet.cell(1,i).value)
        return title


    def do_excel(self):
        #打开工作薄  定位表单
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]

        #获取title
        title = self.get_title()

        all_data = []#大列表
        for j in range (2,sheet.max_row+1):#控制行
            row_data = {}#每一行数据存到一个字典里面
            for i in range(1,sheet.max_column+1):#控制列
                # key = title[i-1]#i=1 是第一列 对应title里面的0    i-1
                # value = sheet.cell(j,i).value
                row_data[title[i-1]]=sheet.cell(j,i).value
            all_data.append(row_data)
        # print(all_data)
        return all_data

if __name__ == '__main__':
    res=DoExcel('testdata.xlsx','my_test_data').do_excel()
    print(res)

