from openpyxl import load_workbook
#请把每一行数据存到字典里面，其中key为对于的表头，所有字典存到一个大列表里面

class DoExcel:
    def __init__(self,file_name,sheet_name):#Excel文件名  表单名
        self.file_name=file_name
        self.sheet_name=sheet_name
        wb = load_workbook(self.file_name)
        self.sheet = wb[self.sheet_name]
    def get_title(self):#获取第一行表头字段
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        title = []  #存储表头
        for i in range(1,sheet.max_column+1):
            title.append(sheet.cell(1,i).value)
        return title
    def do_excel(self,mode,case_id_list): #打开工作薄  定位表单#获取title
        title = self.get_title()
        test_data = []#大列表
        for i in range (2,self.sheet.max_row+1):#控制行
            test_b = {}  # 每一行数据存到一个字典里面
            for j in range(1,self.sheet.max_column+1):#控制列
                test_b[title[j-1]]=self.sheet.cell(i,j).value
            test_data.append(test_b)
        if mode=='1':#如果等于1 就跑所有的用例
            final_data=test_data
        else:#否则就执行case_id_list里面的用例
            final_data = []
            for item in test_data:#对所有的数据进行遍历
                if item['case_id'] in eval(case_id_list):#去判断case_id是否相等
                    final_data.append(item)#如果相等的话，那么就把他加入到final_data的列表里面
        return final_data #返回最终的数据
if __name__ == '__main__':
    from conf1 import read_path1
    r=DoExcel(read_path1.test_data_path,'d1').do_excel('1','[5]')
    print(r)
    print(read_path1.test_data_path)












#
# from openpyxl import load_workbook
# #请把每一行数据存到字典里面，其中key为对于的表头，所有字典存到一个大列表里面
#
# class DoExcel:
#     def __init__(self,file_name,sheet_name):#Excel文件名  表单名
#         self.file_name=file_name
#         self.sheet_name=sheet_name
#
#     def get_title(self):#获取第一行表头字段
#         wb = load_workbook(self.file_name)
#         sheet = wb[self.sheet_name]
#         title = []  #存储表头
#         for i in range(1,sheet.max_column+1):
#             title.append(sheet.cell(1,i).value)
#         return title
#
#
#     def do_excel(self,mode,case_id_list):
#         #打开工作薄  定位表单
#         wb = load_workbook(self.file_name)
#         sheet = wb[self.sheet_name]
#         #获取title
#         title = self.get_title()
#         test_data = []#大列表
#         for j in range (2,sheet.max_row+1):#控制行
#             row_data = {}  # 每一行数据存到一个字典里面
#             for i in range(1,sheet.max_column+1):#控制列
#                 # key = title[i-1]#i=1 是第一列 对应title里面的0    i-1
#                 # value = sheet.cell(j,i).value
#                 row_data[title[i-1]]=sheet.cell(j,i).value
#             test_data.append(row_data)
#             if mode=='1':#如果等于1 就跑所有的用例
#                 final_data=test_data
#             else:#否则就执行case_id_list里面的用例
#                 final_data = []
#                 for item in test_data:#对所有的数据进行遍历
#                     if item['case_id'] in eval(case_id_list):#去判断case_id是否相等
#                         final_data.append(item)#如果相等的话，那么就把他加入到final_data的列表里面
#
#         return final_data #返回最终的数据
#         # return test_data
# if __name__ == '__main__':
#     # from conf1 import read_path
#     path=r'C:\Users\Administrator\Desktop\automation\python9\work2\test_data\t1.xlsx'
#     res=DoExcel(path,'test_data_1').do_excel()
#     print(res)
# #read_path.test_data_path
# #,mode,case_id_list




#
# from openpyxl import load_workbook
# #请把每一行数据存到字典里面，其中key为对于的表头，所有字典存到一个大列表里面
#
# class DoExcel:
#     def __init__(self,file_name,sheet_name):#Excel文件名  表单名
#         self.file_name=file_name
#         self.sheet_name=sheet_name
#
#     def get_title(self):#获取第一行表头字段
#         wb = load_workbook(self.file_name)
#         sheet = wb[self.sheet_name]
#         title = []  #存储表头
#         for i in range(1,sheet.max_column+1):
#             title.append(sheet.cell(1,i).value)
#         return title
#
#
#     def do_excel(self,mode,case_id_list):
#         #打开工作薄  定位表单
#         wb = load_workbook(self.file_name)
#         sheet = wb[self.sheet_name]
#         #获取title
#         title = self.get_title()
#         test_data = []#大列表
#         for j in range (2,sheet.max_row+1):#控制行
#             row_data = {}  # 每一行数据存到一个字典里面
#             for i in range(1,sheet.max_column+1):#控制列
#                 # key = title[i-1]#i=1 是第一列 对应title里面的0    i-1
#                 # value = sheet.cell(j,i).value
#                 row_data[title[i-1]]=sheet.cell(j,i).value
#             test_data.append(row_data)
#             if mode=='1':#如果等于1 就跑所有的用例
#                 final_data=test_data
#             else:#否则就执行case_id_list里面的用例
#                 final_data = []
#                 for item in test_data:#对所有的数据进行遍历
#                     if item['case_id'] in eval(case_id_list):#去判断case_id是否相等
#                         final_data.append(item)#如果相等的话，那么就把他加入到final_data的列表里面
#
#         return final_data #返回最终的数据
#         # return test_data
# if __name__ == '__main__':
#     from conf1 import read_path
#     # path=r'C:\Users\Administrator\Desktop\automation\python9\work2\test_data\t1.xlsx'
#     res=DoExcel(read_path.test_data_path,'test_data_1').do_excel('1','[2,3]')
#     print(res)
# #read_path.test_data_path
# #,mode,case_id_list